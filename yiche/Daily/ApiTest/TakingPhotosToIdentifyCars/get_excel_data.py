# -*- coding: utf-8 -*-
import os
import openpyxl


class Excel:
    def __init__(self, name):
        """找到文件路径"""
        self.filename = os.getcwd() + name

    def open_excel(self):
        """加载excel"""
        open_excel = openpyxl.load_workbook(self.filename)
        return open_excel

    def sheet_names(self):
        """获取所有sheet页名字,列表形式"""
        return self.open_excel().sheetnames

    def sheet(self, index=0):
        """获取sheet页对象,默认第一sheet页"""
        return self.open_excel()[self.sheet_names()[index]]

    def get_data(self, row, column, index=0):
        """获取某一单元格对象"""
        return self.sheet(index).cell(row, column)

    def get_data_value(self, row, column, index=0):
        """获取某一单元格内容"""
        return self.sheet(index).cell(row, column).value

    def get_row(self, row, index=0):
        """获取一行对象"""
        return self.sheet(index)[row]

    def max_row(self, index=0):
        """获取最大行数"""
        return self.sheet(index).max_row

    def max_column(self, index=0):
        """获取最大列数"""
        return self.sheet(index).max_column

    def get_row_value(self, row, index=0):
        """获取一行数据 加入列表"""
        b = []
        for i in self.sheet(index)[row]:
            b.append(i.value)
        return b

    def write_excel(self, row, column, value, index=0):
        wb = self.open_excel()  # 打开excel
        wr = wb[self.sheet_names()[index]]  # 默认读取第一sheet页数据
        wr.cell(row, column, value)  # 写入内容 到第一sheet页某行某列
        wb.save(self.filename)  # 保存


e = Excel("\\NewCarModel.xlsx")

if __name__ == '__main__':
    cars_model = os.listdir(os.getcwd() + '/cars_picture/')
    print(cars_model)
    for row_number in range(e.max_row()-1):
        car_model = e.get_data_value(row_number+2, 6)
        car_id = str(e.get_data_value(row_number+2, 5))
        # os.mkdir(os.getcwd() + '/cars_picture/' + car_model)
        # print(car_id, car_model)
        os.rename(os.getcwd() + '/cars_picture/' + car_model, os.getcwd() + '/cars_picture/' +
                  car_id + '---' + car_model)
